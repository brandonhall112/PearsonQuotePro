import sys, math
from PySide6.QtGui import QDesktopServices
from PySide6.QtCore import QUrl
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Dict as TDict


def resolve_excel_path(expected_name: str = "Tech days and quote rates.xlsx") -> Path | None:
    """Find the default Excel workbook inside assets without prompting the user unless missing."""
    assets = resolve_assets_dir()

    exact = (assets / expected_name)
    try:
        if exact.exists():
            return exact.resolve()
    except Exception:
        pass

    # Fuzzy match (handles minor renames like '(1)' etc.)
    try:
        for f in assets.glob("*.xlsx"):
            n = f.name.lower()
            if "tech" in n and "quote" in n and "rate" in n:
                return f.resolve()
    except Exception:
        pass

    # If exactly one xlsx exists, use it
    try:
        xls = list(assets.glob("*.xlsx"))
        if len(xls) == 1:
            return xls[0].resolve()
    except Exception:
        pass

    return None



def resolve_assets_dir() -> Path:
    """Return the assets directory for dev + PyInstaller (onefile/onedir).

    onefile: extracted to sys._MEIPASS/assets
    onedir: typically <exe_dir>/_internal/assets (new PyInstaller layout) or <exe_dir>/assets
    dev:   <repo>/assets
    """
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        return (Path(meipass).resolve() / "assets").resolve()

    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        for p in (exe_dir / "_internal" / "assets", exe_dir / "assets"):
            try:
                if p.exists():
                    return p.resolve()
            except Exception:
                pass
        return (exe_dir / "_internal" / "assets").resolve()

    return (Path(__file__).resolve().parent / "assets").resolve()


import numpy as np
import openpyxl

from PySide6.QtCore import Qt, QSize
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QSpinBox,
    QComboBox, QCheckBox, QFrame, QScrollArea, QSplitter,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QSizePolicy
)
from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
from PySide6.QtGui import QTextDocument
from PySide6.QtGui import QPageSize, QFont, QPainter, QColor
from PySide6.QtCharts import QChart, QChartView, QHorizontalBarSeries, QHorizontalStackedBarSeries, QBarSet, QValueAxis, QBarCategoryAxis
import base64

APP_TITLE = "Pearson Commissioning Pro"

# Business rules
TRAINING_MACHINES_PER_DAY = 3  # 1 training day per 3 machines (ceil)
DEFAULT_INSTALL_WINDOW = 7
MIN_INSTALL_WINDOW = 3
MAX_INSTALL_WINDOW = 14
TRAVEL_DAYS_PER_PERSON = 2  # travel-in + travel-out

# Requested overrides
OVERRIDE_AIRFARE_PER_PERSON = 1500.0
OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON = 150.0

ASSETS_DIR = resolve_assets_dir()
DEFAULT_EXCEL = resolve_excel_path() or (ASSETS_DIR / "Tech days and quote rates.xlsx")
LOGO_PATH = ASSETS_DIR / "Pearson Logo.png"


def ceil_int(x: float) -> int:
    return int(math.ceil(float(x)))


def balanced_allocate(total_days: int, headcount: int) -> List[int]:
    """Balance integer days to minimize the maximum assigned days."""
    if headcount <= 0:
        return []
    loads = [0] * headcount
    for _ in range(int(total_days)):
        i = int(np.argmin(loads))
        loads[i] += 1
    loads.sort(reverse=True)
    return loads




def chunk_allocate_by_machine(install_days_per_machine: int, qty: int, training_days: int, window: int) -> List[int]:
    """Allocate work using whole-machine install chunks + whole-day training chunks.

    Install days are assigned per machine (no fractional splitting). Training days are 1-day chunks.
    We choose the *minimum* headcount that keeps every person's onsite days <= window.

    Training assignment heuristic:
      - Prefer assigning training to the currently *most-loaded* person that can still accept a day
        without exceeding the window (keeps extra people from traveling and mirrors reality).
      - If none can accept, fall back to the least-loaded person (best-effort).
    Returns a list of total onsite days per person (sorted descending).
    """
    install_days_per_machine = int(install_days_per_machine or 0)
    qty = int(qty or 0)
    training_days = int(training_days or 0)
    window = int(window or 0)

    if window <= 0:
        return []
    if qty <= 0 and training_days <= 0:
        return []

    # If there is no install work, allocate training only.
    if qty <= 0 or install_days_per_machine <= 0:
        headcount = ceil_int(training_days / window) if training_days > 0 else 0
        loads = balanced_allocate(training_days, headcount) if headcount > 0 else []
        return loads

    max_headcount = max(1, qty)  # at most one machine per person
    for headcount in range(1, max_headcount + 1):
        # Distribute whole machines as evenly as possible.
        base_n = qty // headcount
        rem = qty % headcount
        machine_counts = [base_n + (1 if i < rem else 0) for i in range(headcount)]
        loads = [c * install_days_per_machine for c in machine_counts]

        # Assign training days as 1-day chunks.
        for _ in range(training_days):
            # Prefer adding to the most-loaded person who can still accept 1 day within window.
            candidates = [i for i, d in enumerate(loads) if d + 1 <= window]
            if candidates:
                i = max(candidates, key=lambda j: loads[j])
            else:
                i = int(np.argmin(loads))
            loads[i] += 1

        if max(loads) <= window:
            loads.sort(reverse=True)
            return loads

    # Best-effort fallback (should generally be prevented by validation).
    loads = [install_days_per_machine] * qty
    for _ in range(training_days):
        i = int(np.argmin(loads))
        loads[i] += 1
    loads.sort(reverse=True)
    return loads

@dataclass
class ModelInfo:
    item: str
    tech_install_days_per_machine: int   # install-only (no training baked in)
    eng_days_per_machine: int
    training_applicable: bool = True


@dataclass
class LineSelection:
    model: str
    qty: int
    training_required: bool


@dataclass
class RoleTotals:
    headcount: int
    total_onsite_days: int
    onsite_days_by_person: List[int]
    day_rate: float
    labor_cost: float


@dataclass
class ExpenseLine:
    description: str
    quantity: float
    unit_price: float
    extended: float
    details: str


@dataclass
class Assignment:
    model: str
    role: str  # "Technician" or "Engineer"
    person_num: int
    onsite_days: int
    cost: float


class ExcelData:
    def __init__(self, path: Path):
        self.path = path
        self.models: Dict[str, ModelInfo] = {}
        self.rates: Dict[str, Dict[str, object]] = {}
        self.requirements: List[str] = []
        self._load()

    def _load(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)

        # Models: Instal days by Model
        if "Instal days by Model" not in wb.sheetnames:
            raise ValueError("Missing sheet: 'Instal days by Model'")
        ws = wb["Instal days by Model"]

        headers = {
            str(ws.cell(1, c).value).strip(): c
            for c in range(1, ws.max_column + 1)
            if ws.cell(1, c).value is not None
        }

        def find_col(pred):
            for k, c in headers.items():
                if pred(k.lower()):
                    return c
            return None

        col_item = find_col(lambda s: s in ["item", "model", "machine", "machine type"])
        col_tech = find_col(lambda s: "technician" in s and "day" in s)
        col_eng = find_col(lambda s: ("engineer" in s and "day" in s) or ("field engineer" in s and "day" in s))
        col_train_app = find_col(lambda s: ("training required" in s))

        if col_item is None or col_tech is None or col_eng is None:
            raise ValueError("Model sheet columns not found. Expected: Item, Technician Days Required, Field Engineer Days Required.")

        def _as_bool(v, default=True):
            if v is None:
                return default
            if isinstance(v, bool):
                return v
            s = str(v).strip().lower()
            if s in ("true","t","yes","y","1"):
                return True
            if s in ("false","f","no","n","0"):
                return False
            return default

        for r in range(2, ws.max_row + 1):
            item = ws.cell(r, col_item).value
            if item is None:
                continue
            item = str(item).strip()
            if not item:
                continue
            tech = ws.cell(r, col_tech).value or 0
            eng = ws.cell(r, col_eng).value or 0
            try:
                tech_i = int(float(tech))
            except Exception:
                tech_i = 0
            try:
                eng_i = int(float(eng))
            except Exception:
                eng_i = 0
            train_app = _as_bool(ws.cell(r, col_train_app).value, default=True) if col_train_app is not None else True
            self.models[item] = ModelInfo(item=item, tech_install_days_per_machine=tech_i, eng_days_per_machine=eng_i, training_applicable=train_app)

        # Rates: Service Rates
        if "Service Rates" not in wb.sheetnames:
            raise ValueError("Missing sheet: 'Service Rates'")
        ws = wb["Service Rates"]

        header_row = None
        for r in range(1, 15):
            if str(ws.cell(r, 2).value).strip().lower() == "item" and str(ws.cell(r, 3).value).strip().lower() == "description":
                header_row = r
                break
        if header_row is None:
            header_row = 3

        for r in range(header_row + 1, ws.max_row + 1):
            desc = ws.cell(r, 3).value
            if desc is None:
                continue
            desc_s = str(desc).strip()
            if not desc_s:
                continue
            unit = ws.cell(r, 6).value
            notes = ws.cell(r, 7).value
            try:
                unit_f = float(unit)
            except Exception:
                continue
            self.rates[desc_s.lower()] = {
                "description": desc_s,
                "unit_price": unit_f,
                "notes": str(notes).strip() if notes is not None else ""
            }

        # Requirements & Assumptions
        if "Requirements and Assumptions" in wb.sheetnames:
            ws = wb["Requirements and Assumptions"]
            out = []
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r, 3).value
                if v is None:
                    continue
                s = str(v).strip()
                if s and not s.lower().startswith("assumptions and requirements"):
                    out.append(s)
            self.requirements = out

    def get_rate(self, key: str) -> Tuple[float, str]:
        k = key.lower().strip()
        if k in self.rates:
            rv = self.rates[k]
            return float(rv["unit_price"]), str(rv["description"])
        for rk, rv in self.rates.items():
            if k in rk:
                return float(rv["unit_price"]), str(rv["description"])
        raise KeyError(f"Rate not found for '{key}'")


class MachineLine(QFrame):
    def __init__(self, models: List[str], training_applicable_map: Dict[str, bool], on_change, on_delete):
        super().__init__()
        self.on_change = on_change
        self.on_delete = on_delete
        self.training_applicable_map = training_applicable_map

        self.setObjectName("machineLine")
        row = QHBoxLayout(self)
        row.setContentsMargins(10, 10, 10, 10)
        row.setSpacing(10)

        self.cmb_model = QComboBox()
        self.cmb_model.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.cmb_model.addItem("— Select —")
        self.cmb_model.addItems(models)
        self.cmb_model.setCurrentIndex(-1)
        self.cmb_model.currentIndexChanged.connect(self._model_changed)

        self.spin_qty = QSpinBox()
        self.spin_qty.setRange(0, 999)
        self.spin_qty.setValue(1)
        self.spin_qty.valueChanged.connect(self._changed)

        self.chk_training = QCheckBox("Training Required")
        self.chk_training.setChecked(True)
        self.chk_training.stateChanged.connect(self._changed)
        self.chk_training.setToolTip("Uncheck only by customer request (training is normally included).")

        self.btn_delete = QPushButton("🗑")
        self.btn_delete.setFixedWidth(40)
        self.btn_delete.clicked.connect(self._delete)

        row.addWidget(QLabel("Machine Model"))
        row.addWidget(self.cmb_model, 2)
        row.addWidget(QLabel("Qty"))
        row.addWidget(self.spin_qty)
        row.addWidget(self.chk_training, 1)
        row.addWidget(self.btn_delete)

        self._model_changed()
    def _model_changed(self, *_):
        model = self.cmb_model.currentText().strip()
        if model == "— Select —":
            model = ""
        if not model:
            self.chk_training.hide()
            self.chk_training.setChecked(False)
        else:
            applicable = bool(self.training_applicable_map.get(model, True))
            if not applicable:
                self.chk_training.hide()
                self.chk_training.setChecked(False)
            else:
                self.chk_training.show()
                if not self.chk_training.isChecked():
                    self.chk_training.setChecked(True)
        self.on_change()

    def _changed(self, *_):

        self.on_change()

    def _delete(self):
        self.on_delete(self)

    def value(self) -> LineSelection:
        model = self.cmb_model.currentText().strip()
        if model == "— Select —":
            model = ""
        return LineSelection(
            model=model,
            qty=int(self.spin_qty.value()) if model else 0,
            training_required=(bool(self.chk_training.isChecked()) if self.chk_training.isVisible() else False)
        )


class Card(QFrame):
    def __init__(self, title: str, icon_text: str):
        super().__init__()
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setObjectName("card")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)

        top = QHBoxLayout()
        ic = QLabel(icon_text)
        ic.setObjectName("cardIcon")
        ic.setFixedSize(QSize(34, 34))
        ic.setAlignment(Qt.AlignCenter)
        top.addWidget(ic)

        v = QVBoxLayout()
        self.lbl_title = QLabel(title)
        self.lbl_title.setObjectName("cardTitle")
        self.lbl_value = QLabel("—")
        self.lbl_value.setObjectName("cardValue")
        self.lbl_sub = QLabel("")
        self.lbl_sub.setObjectName("cardSub")
        self.lbl_sub.setWordWrap(True)
        v.addWidget(self.lbl_title)
        v.addWidget(self.lbl_value)
        top.addLayout(v, 1)

        lay.addLayout(top)
        lay.addWidget(self.lbl_sub)

    def set_value(self, value: str, sub: str = ""):
        self.lbl_value.setText(value)
        self.lbl_sub.setText(sub or "")


class Section(QFrame):
    def __init__(self, title: str, subtitle: str = "", icon_text: str = ""):
        super().__init__()
        self.setObjectName("section")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 14, 14, 14)
        lay.setSpacing(10)

        head = QHBoxLayout()
        if icon_text:
            ic = QLabel(icon_text)
            ic.setObjectName("sectionIcon")
            ic.setFixedSize(QSize(28, 28))
            ic.setAlignment(Qt.AlignCenter)
            head.addWidget(ic)

        title_box = QVBoxLayout()
        t = QLabel(title)
        t.setObjectName("sectionTitle")
        title_box.addWidget(t)
        if subtitle:
            s = QLabel(subtitle)
            s.setObjectName("sectionSub")
            s.setWordWrap(True)
            title_box.addWidget(s)
        head.addLayout(title_box, 1)
        lay.addLayout(head)

        self.content = QWidget()
        self.content_layout = QVBoxLayout(self.content)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(10)
        lay.addWidget(self.content)


def money(x: float) -> str:
    return f"${x:,.0f}"


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(APP_TITLE)
        self.resize(1920, 1200)

        if not DEFAULT_EXCEL or not Path(DEFAULT_EXCEL).exists():
            raise FileNotFoundError("Missing required workbook: Tech days and quote rates.xlsx")
        self.data = ExcelData(Path(DEFAULT_EXCEL))
        self.models_sorted = sorted(self.data.models.keys())
        self.training_app_map = {k: bool(v.training_applicable) for k, v in self.data.models.items()}
        self.lines: List[MachineLine] = []

        central_container = QWidget()
        root = QVBoxLayout(central_container)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Outer scroll area enables whole-window scrolling in stacked (single-column) mode
        self.outer_scroll = QScrollArea()
        self.outer_scroll.setObjectName("outerScroll")
        self.outer_scroll.setWidgetResizable(True)
        self.outer_scroll.setFrameShape(QFrame.NoFrame)
        self.outer_scroll.setWidget(central_container)
        self.setCentralWidget(self.outer_scroll)


        header = QFrame()
        header.setObjectName("header")
        h = QHBoxLayout(header)
        h.setContentsMargins(14, 10, 14, 10)
        self.lbl_title = QLabel("🧾  " + APP_TITLE)
        self.lbl_title.setObjectName("appTitle")
        h.addWidget(self.lbl_title)
        h.addStretch(1)
        btn_load = QPushButton("Load Excel…")
        btn_load.setToolTip("Load a different Excel workbook (will replace the bundled rates/models for this session).")
        btn_load.clicked.connect(self.open_excel)
        h.addWidget(btn_load)

        btn_open_bundled = QPushButton("Open Bundled Excel")
        btn_open_bundled.setToolTip("Open the Excel workbook that was bundled into this EXE (for verification).")
        btn_open_bundled.clicked.connect(self.open_bundled_excel)
        h.addWidget(btn_open_bundled)
        root.addWidget(header)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        self.splitter = splitter
        root.addWidget(splitter, 1)

        # LEFT
        left = QFrame()
        left.setObjectName("panel")
        left_l = QVBoxLayout(left)
        left_l.setContentsMargins(14, 14, 14, 14)
        left_l.setSpacing(12)

        t = QLabel("Machine Configuration")
        t.setObjectName("panelTitle")
        left_l.addWidget(t)

        left_l.addWidget(QLabel(
            "Add machines to estimate commissioning requirements.\\n"
            "Each machine type requires dedicated personnel (no sharing across types)."
        ))

        win_box = QFrame()
        win_box.setObjectName("softBox")
        win_l = QHBoxLayout(win_box)
        win_l.setContentsMargins(12, 10, 12, 10)
        win_l.addWidget(QLabel("Customer Install Window"))
        self.spin_window = QSpinBox()
        self.spin_window.setRange(MIN_INSTALL_WINDOW, MAX_INSTALL_WINDOW)
        self.spin_window.setValue(DEFAULT_INSTALL_WINDOW)
        self.spin_window.valueChanged.connect(self.recalc)
        win_l.addStretch(1)
        win_l.addWidget(self.spin_window)
        win_l.addWidget(QLabel("days"))
        left_l.addWidget(win_box)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        container = QWidget()
        self.lines_layout = QVBoxLayout(container)
        self.lines_layout.setContentsMargins(0, 0, 0, 0)
        self.lines_layout.setSpacing(10)

        self.empty_hint = QLabel("No machines added.\\nClick “Add Machine” to begin.")
        self.empty_hint.setObjectName("emptyHint")
        self.empty_hint.setAlignment(Qt.AlignCenter)
        self.empty_hint.setMinimumHeight(120)
        self.lines_layout.addWidget(self.empty_hint)

        self.scroll.setWidget(container)
        left_l.addWidget(self.scroll, 1)

        btn_add = QPushButton("+  Add Machine")
        btn_add.setObjectName("addMachine")
        btn_add.clicked.connect(self.add_line)
        left_l.addWidget(btn_add)

        note = QLabel("Note: Unchecking “Training Required” should only be done by customer request.")
        note.setObjectName("note")
        note.setWordWrap(True)
        left_l.addWidget(note)

        splitter.addWidget(left)

        # RIGHT (scrollable)
        right_wrap = QWidget()
        self.right_wrap = right_wrap
        right_layout = QVBoxLayout(right_wrap)
        right_layout.setContentsMargins(0, 0, 0, 0)

        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        self.right_scroll = right_scroll
        right_layout.addWidget(right_scroll)

        right = QWidget()
        self.right_content = right
        right_scroll.setWidget(right)
        right_l = QVBoxLayout(right)
        right_l.setContentsMargins(14, 14, 14, 14)
        right_l.setSpacing(12)

        cards = QHBoxLayout()
        self.card_tech = Card("Technicians", "🧰")
        self.card_eng = Card("Engineers", "🧑‍💻")
        self.card_window = Card("Max Onsite", "⏱")
        self.card_total = Card("Total Cost", "💲")
        cards.addWidget(self.card_tech, 1)
        cards.addWidget(self.card_eng, 1)
        cards.addWidget(self.card_window, 1)
        cards.addWidget(self.card_total, 1)
        right_l.addLayout(cards)

        self.alert = QLabel("")
        self.alert.setObjectName("alert")
        self.alert.setWordWrap(True)
        self.alert.hide()
        right_l.addWidget(self.alert)

        self.tbl_breakdown = self.make_table(["Model", "Qty", "Tech Days", "Eng Days", "Technicians", "Engineers"])
        self.tbl_assign = self.make_table(["Machine Type", "Role", "Person #", "Assigned Days", "Cost"])
        self.tbl_labor = self.make_table(["Role", "Daily Rate", "Total Days", "Personnel", "Total Cost"])
        self.tbl_exp = self.make_table(["Expense", "Details", "Amount"])
        self.tbl_exp.setMinimumHeight(0)

        sec_breakdown = Section("Machine Breakdown", "Days and personnel required per machine model", "🧩")
        sec_breakdown.content_layout.addWidget(self.tbl_breakdown)

        sec_assign = Section("Personnel Assignments", "Each machine type has dedicated personnel.", "👥")
        sec_assign.content_layout.addWidget(self.tbl_assign)

        sec_labor = Section("Labor Costs", "Labor costs by role at daily rates (8 hours/day).", "🛠")
        sec_labor.content_layout.addWidget(self.tbl_labor)

        # Workload bar chart (bonus visual)
        self.chart = QChart()
        self.chart_view = QChartView(self.chart)
        self.chart_view.setRenderHint(QPainter.Antialiasing)
        self.chart_view.setMinimumHeight(300)
        sec_chart = Section("Workload", "Days onsite per person (T=Tech, E=Engineer).", "📊")
        sec_chart.content_layout.addWidget(self.chart_view)

        # Left side: put chart under Machine Configuration so the right-side widgets stay readable
        left_l.addWidget(sec_chart)

        right_l.addWidget(sec_breakdown)
        right_l.addWidget(sec_assign)
        right_l.addWidget(sec_labor)

        sec_exp = Section("Estimated Expenses", "", "🧳")
        self.lbl_exp_hdr = QLabel("")
        self.lbl_exp_hdr.setObjectName("sectionSub")
        self.lbl_exp_hdr.setWordWrap(True)
        sec_exp.content_layout.addWidget(self.lbl_exp_hdr)
        sec_exp.content_layout.addWidget(self.tbl_exp)
        right_l.addWidget(sec_exp)

        bottom = QFrame()
        bottom.setObjectName("totalBar")
        bl = QHBoxLayout(bottom)
        bl.setContentsMargins(14, 12, 14, 12)
        self.lbl_total = QLabel("Estimated Total")
        self.lbl_total.setObjectName("totalLabel")
        self.lbl_total_val = QLabel("—")
        self.lbl_total_val.setObjectName("totalValue")
        bl.addWidget(self.lbl_total)
        bl.addStretch(1)
        bl.addWidget(self.lbl_total_val)
        self.btn_print = QPushButton("Print Quote…")
        self.btn_print.clicked.connect(self.print_quote_preview)
        self.btn_print.setEnabled(False)
        bl.addWidget(self.btn_print)
        right_l.addWidget(bottom)

        splitter.addWidget(right_wrap)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)

        self.apply_theme()
        self.reset_views()

        # Responsive scaling baseline (designed for 1920x1200)
        self._base_font_pt = float(self.font().pointSizeF() or 10.0)
        self._apply_scale()

        # Responsive layout: two-column w/ right scroll on large screens; single stacked w/ full-window scroll on small screens
        self._stack_threshold = 1280
        self._is_stacked = False
        self._apply_responsive_layout()

    def make_table(self, headers: List[str]) -> QTableWidget:
        tbl = QTableWidget(0, len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl.setSelectionMode(QAbstractItemView.SingleSelection)
        tbl.setAlternatingRowColors(True)
        tbl.setObjectName("table")
        tbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        tbl.setMinimumHeight(120)
        return tbl

    def apply_theme(self):
        # Pearson-ish palette (navy + orange + neutral)
        blue = "#4B4F54"   # charcoal gray (logo text)
        gold = "#F05A28"   # Pearson orange
        neutral = "#6D6E71"
        red = "#D6453D"    # Pearson red accent
        css = """
        QFrame#header { background: __BLUE__; color: white; border: none; }
        QLabel#appTitle { color: white; font-size: 20px; font-weight: 800; }
        QFrame#panel { background: white; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#panelTitle { font-size: 16px; font-weight: 800; color: #0F172A; }
        QFrame#softBox { background: #FFF7EA; border: 1px solid #F0D8A8; border-radius: 12px; }
        QLabel#note { color: #334155; font-size: 12px; }
        QLabel#emptyHint { color: __NEUTRAL__; background: #F8FAFC; border: 1px dashed #CBD5E1; border-radius: 12px; }
        QPushButton#primary {
            background: __GOLD__; border: 0px; color: #0B1B2A;
            padding: 10px 12px; border-radius: 10px; font-weight: 800;
        }
        QPushButton#addMachine {
            background: #bebebe; border: 0px; color: #0B1B2A;
            padding: 10px 12px; border-radius: 10px; font-weight: 800;
        }
        QPushButton#addMachine:hover { background: #D6D9DD; }
        QPushButton#addMachine:pressed { background: #CBD5E1; }
        QPushButton {
            padding: 8px 10px; border-radius: 10px;
            border: 1px solid #D6D9DD; background: #F8FAFC;
        }
        QPushButton:disabled { color: #94A3B8; background: #F1F5F9; }
        QFrame#card { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#cardIcon { background: #EEF2F7; border-radius: 10px; font-size: 16px; }
        QLabel#cardTitle { font-size: 12px; color: __NEUTRAL__; font-weight: 700; }
        QLabel#cardValue { font-size: 24px; color: #0F172A; font-weight: 900; }
        QLabel#cardSub { font-size: 12px; color: __NEUTRAL__; }
        QFrame#section { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#sectionIcon { background: #EEF2F7; border-radius: 10px; font-size: 14px; }
        QLabel#sectionTitle { font-size: 15px; font-weight: 900; color: #0F172A; }
        QLabel#sectionSub { font-size: 12px; color: __NEUTRAL__; }
        QTableWidget#table {
            background: #FFFFFF;
            border: 1px solid #EEF0F2;
            border-radius: 12px;
            gridline-color: #E2E8F0;
            selection-background-color: #DBEAFE;
        }
        QHeaderView::section {
            background: #343551;
            color: white;
            padding: 8px;
            border: 0px;
            border-bottom: 1px solid #E2E8F0;
            font-weight: 800;
        }
        QFrame#totalBar { background: #F8FAFC; border: 1px solid #E6E8EB; border-radius: 14px; }
        QLabel#totalLabel { font-size: 13px; font-weight: 800; color: #0F172A; }
        QLabel#totalValue { font-size: 22px; font-weight: 900; color: __BLUE__; }
        QLabel#alert {
            background: #FEF2F2;
            border: 1px solid #FCA5A5;
            color: #7F1D1D;
            padding: 10px;
            border-radius: 12px;
        }
        QFrame#machineLine { background: #FFFFFF; border: 1px solid #E6E8EB; border-radius: 12px; }
        """

        css = css.replace("__BLUE__", blue).replace("__GOLD__", gold).replace("__NEUTRAL__", neutral).replace("__RED__", red)
        self.setStyleSheet(css)

    def reset_views(self):
        self.card_tech.set_value("0", "0 total days")
        self.card_eng.set_value("0", "0 total days")
        self.card_window.set_value(f"{self.spin_window.value()}", "")
        self.card_total.set_value("—", "labor + expenses")
        self.lbl_total_val.setText("—")
        for tbl in [self.tbl_breakdown, self.tbl_assign, self.tbl_labor, self.tbl_exp]:
            tbl.setRowCount(0)
        self.lbl_exp_hdr.setText("")
        self.btn_print.setEnabled(False)
        self.alert.hide()
        self.alert.setText("")
        if hasattr(self, 'chart'):
            try:
                self.chart.removeAllSeries()
                self.chart.setTitle('Workload (onsite days)')
            except Exception:
                pass

    def add_line(self):
        if self.empty_hint is not None:
            self.empty_hint.hide()
        ln = MachineLine(self.models_sorted, self.training_app_map, on_change=self.recalc, on_delete=self.delete_line)
        self.lines.append(ln)
        self.lines_layout.addWidget(ln)
        self.recalc()

    def delete_line(self, ln: MachineLine):
        self.lines.remove(ln)
        ln.setParent(None)
        ln.deleteLater()
        if len(self.lines) == 0:
            self.empty_hint.show()
            self.reset_views()
        else:
            self.recalc()

    def open_bundled_excel(self):
        """Open the bundled Excel workbook in the user's default spreadsheet app."""
        try:
            p = resolve_excel_path()
            if not p or not p.exists():
                QMessageBox.warning(self, "Bundled Excel not found",
                                    "The bundled Excel file could not be found inside the app assets.")
                return
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(p)))
        except Exception as e:
            QMessageBox.critical(self, "Open error", str(e))


    def open_excel(self):
        fp, _ = QFileDialog.getOpenFileName(self, "Select Excel file", "", "Excel (*.xlsx)")
        if not fp:
            return
        try:
            self.data = ExcelData(Path(fp))
            self.models_sorted = sorted(self.data.models.keys())
            for ln in self.lines:
                cur = ln.cmb_model.currentText()
                ln.cmb_model.blockSignals(True)
                ln.cmb_model.clear()
                ln.cmb_model.addItems(self.models_sorted)
                if cur in self.models_sorted:
                    ln.cmb_model.setCurrentIndex(self.models_sorted.index(cur))
                ln.cmb_model.blockSignals(False)
            self.recalc()
        except Exception as e:
            QMessageBox.critical(self, "Excel load error", str(e))

    def calc(self):
        selections = [ln.value() for ln in self.lines]
        selections = [s for s in selections if s.qty > 0 and s.model and s.model in self.data.models]
        if not selections:
            raise ValueError("No machines selected. Click “Add Machine” to begin.")

        window = int(self.spin_window.value())

        tech_hr, _ = self.data.get_rate("tech. regular time")
        eng_hr, _ = self.data.get_rate("eng. regular time")
        hours_per_day = 8
        tech_day_rate = tech_hr * hours_per_day
        eng_day_rate = eng_hr * hours_per_day

        machine_rows = []
        assignments: List[Assignment] = []
        tech_all: List[int] = []
        eng_all: List[int] = []

        for s in selections:
            mi = self.data.models[s.model]
            base_training = ceil_int(s.qty / TRAINING_MACHINES_PER_DAY) if mi.training_applicable else 0
            training_days = base_training if s.training_required else 0

            tech_install_total = mi.tech_install_days_per_machine * s.qty
            tech_total = tech_install_total + training_days
            eng_training_potential = base_training if (mi.eng_days_per_machine > 0) else 0
            eng_training_days = eng_training_potential if s.training_required else 0
            eng_total = (mi.eng_days_per_machine * s.qty) + eng_training_days

            single_training = 1 if (s.training_required and mi.training_applicable) else 0
            if mi.tech_install_days_per_machine + single_training > window:
                raise ValueError(f"{s.model}: Install ({mi.tech_install_days_per_machine}) + Training ({single_training}) exceeds the Customer Install Window ({window}).")
            if mi.eng_days_per_machine > 0:
                single_eng_training = 1 if (s.training_required and mi.eng_days_per_machine > 0) else 0
                if mi.eng_days_per_machine + single_eng_training > window:
                    raise ValueError(
                        f"{s.model}: Engineer ({mi.eng_days_per_machine}) + Training ({single_eng_training}) exceeds the Customer Install Window ({window})."
                    )

            tech_headcount = 0
            eng_headcount = 0

            if tech_total > 0:
                tech_alloc = chunk_allocate_by_machine(mi.tech_install_days_per_machine, s.qty, training_days, window)
                tech_headcount = len(tech_alloc)
                tech_all.extend(tech_alloc)
                for i, d in enumerate(tech_alloc, 1):
                    assignments.append(Assignment(s.model, "Technician", i, d, d * tech_day_rate))

            if eng_total > 0:
                eng_alloc = chunk_allocate_by_machine(mi.eng_days_per_machine, s.qty, eng_training_days, window)
                eng_headcount = len(eng_alloc)
                eng_all.extend(eng_alloc)
                for i, d in enumerate(eng_alloc, 1):
                    assignments.append(Assignment(s.model, "Engineer", i, d, d * eng_day_rate))

            machine_rows.append({
                "model": s.model,
                "qty": s.qty,
                "training_days": training_days,
                "training_potential": base_training,
                "training_required": s.training_required,
                "training_applicable": bool(mi.training_applicable),
                "eng_training_days": eng_training_days,
                "eng_training_potential": eng_training_potential,
                "tech_total": tech_total,
                "eng_total": eng_total,
                "tech_headcount": tech_headcount,
                "eng_headcount": eng_headcount
            })

        tech = RoleTotals(len(tech_all), sum(tech_all), sorted(tech_all, reverse=True), tech_day_rate, float(sum(tech_all)) * tech_day_rate)
        eng = RoleTotals(len(eng_all), sum(eng_all), sorted(eng_all, reverse=True), eng_day_rate, float(sum(eng_all)) * eng_day_rate)

        trip_days_by_person = [a.onsite_days + TRAVEL_DAYS_PER_PERSON for a in assignments]
        n_people = len(trip_days_by_person)
        total_trip_days = sum(trip_days_by_person)
        total_hotel_nights = sum(max(d - 1, 0) for d in trip_days_by_person)

        exp_lines: List[ExpenseLine] = []

        def add_exp(name, qty, unit, detail):
            exp_lines.append(ExpenseLine(name, float(qty), float(unit), float(qty) * float(unit), detail))

        add_exp("Airfare", n_people, OVERRIDE_AIRFARE_PER_PERSON, f"{n_people} person(s) × {money(OVERRIDE_AIRFARE_PER_PERSON)}")
        add_exp("Baggage", total_trip_days, OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON, f"{int(total_trip_days)} day(s) × {money(OVERRIDE_BAGGAGE_PER_DAY_PER_PERSON)}")

        parking, _ = self.data.get_rate("parking")
        car, _ = self.data.get_rate("car rental")
        hotel, _ = self.data.get_rate("hotel")
        per_diem, _ = self.data.get_rate("per diem weekday")
        prep, _ = self.data.get_rate("pre/post trip prep")
        travel_time_rate, _ = self.data.get_rate("travel time")

        add_exp("Car Rental", total_trip_days, car, f"{int(total_trip_days)} day(s) × {money(car)}")
        add_exp("Parking", total_trip_days, parking, f"{int(total_trip_days)} day(s) × {money(parking)}")
        add_exp("Hotel", total_hotel_nights, hotel, f"{int(total_hotel_nights)} night(s) × {money(hotel)}")
        add_exp("Per Diem", total_trip_days, per_diem, f"{int(total_trip_days)} day(s) × {money(per_diem)}")
        add_exp("Pre/Post Trip Prep", n_people, prep, f"{n_people} person(s) × {money(prep)}")
        travel_hours = 16 * n_people
        add_exp("Travel Time", travel_hours, travel_time_rate, f"{travel_hours} hr(s) × {money(travel_time_rate)}/hr")

        exp_total = sum(l.extended for l in exp_lines)
        max_onsite = max([a.onsite_days for a in assignments], default=0)
        grand_total = exp_total + tech.labor_cost + eng.labor_cost

        meta = {
            "machine_rows": machine_rows,
            "assignments": assignments,
            "window": window,
            "max_onsite": max_onsite,
            "n_people": n_people,
            "total_trip_days": total_trip_days,
            "exp_total": exp_total,
            "grand_total": grand_total
        }
        return tech, eng, exp_lines, meta


    def _autosize_table_height(self, tbl, visible_rows=None, max_height=520):
        """Resize table height to fit contents (optionally cap by visible row count) to avoid inner scrolling."""
        try:
            tbl.resizeRowsToContents()
            header_h = tbl.horizontalHeader().height()
            frame = tbl.frameWidth() * 2
            total = header_h + frame + 12
            n = tbl.rowCount()
            if visible_rows is not None:
                n = min(n, int(visible_rows))
            for r in range(n):
                total += tbl.rowHeight(r)
            total = min(total, max_height)
            tbl.setMinimumHeight(total)
            tbl.setMaximumHeight(total)
        except Exception:
            pass


    
    
    def update_workload_chart(self, tech: RoleTotals, eng: RoleTotals):
        """Polished horizontal stacked bar chart of onsite + travel days by person."""
        labels: List[str] = []
        tech_vals: List[int] = []
        eng_vals: List[int] = []

        for d in tech.onsite_days_by_person:
            labels.append(f"T{len(tech_vals)+1}")
            tech_vals.append(int(d))

        for d in eng.onsite_days_by_person:
            labels.append(f"E{len(eng_vals)+1}")
            eng_vals.append(int(d))

        self.chart.removeAllSeries()
        self.chart.setTitle("Workload (days)")
        self.chart.setBackgroundRoundness(8)
        self.chart.setAnimationOptions(QChart.SeriesAnimations)

        if len(labels) == 0:
            return

        # Colors (match UI theme)
        tech_color = QColor("#e04426")  # Tech bar
        eng_color = QColor("#6790a0")   # Engineer bar
        tech_travel = QColor(tech_color); tech_travel.setAlpha(110)
        eng_travel = QColor(eng_color); eng_travel.setAlpha(110)

        series = QHorizontalStackedBarSeries()

        set_tech_on = QBarSet("Tech")
        set_tech_tr = QBarSet("Tech travel")
        set_eng_on = QBarSet("Eng")
        set_eng_tr = QBarSet("Eng travel")

        set_tech_on.setColor(tech_color)
        set_tech_tr.setColor(tech_travel)
        set_eng_on.setColor(eng_color)
        set_eng_tr.setColor(eng_travel)

        n = len(labels)
        # Build arrays aligned to labels: first tech people then engineer people
        for i in range(n):
            is_tech = labels[i].startswith("T")
            if is_tech:
                v = tech_vals[int(labels[i][1:]) - 1]
                set_tech_on.append(float(v))
                set_tech_tr.append(float(TRAVEL_DAYS_PER_PERSON) if v > 0 else 0.0)
                set_eng_on.append(0.0)
                set_eng_tr.append(0.0)
            else:
                v = eng_vals[int(labels[i][1:]) - 1]
                set_tech_on.append(0.0)
                set_tech_tr.append(0.0)
                set_eng_on.append(float(v))
                set_eng_tr.append(float(TRAVEL_DAYS_PER_PERSON) if v > 0 else 0.0)

        series.append(set_tech_on)
        series.append(set_tech_tr)
        series.append(set_eng_on)
        series.append(set_eng_tr)

        self.chart.addSeries(series)

        axis_y = QBarCategoryAxis()
        axis_y.append(labels)

        totals = []
        for i in range(n):
            if labels[i].startswith("T"):
                v = tech_vals[int(labels[i][1:]) - 1]
            else:
                v = eng_vals[int(labels[i][1:]) - 1]
            totals.append(v + (TRAVEL_DAYS_PER_PERSON if v > 0 else 0))
        max_v = max(totals) if totals else 1

        axis_x = QValueAxis()
        axis_x.setRange(0, max(1, int(max_v)))
        axis_x.setLabelFormat("%d")
        axis_x.setTickCount(min(10, max(2, int(max_v) + 1)))

        for ax in list(self.chart.axes()):
            self.chart.removeAxis(ax)

        self.chart.addAxis(axis_y, Qt.AlignLeft)
        self.chart.addAxis(axis_x, Qt.AlignBottom)
        series.attachAxis(axis_y)
        series.attachAxis(axis_x)

        # Labels/legend polish
        try:
            series.setLabelsVisible(True)
            series.setLabelsPosition(series.LabelsInsideEnd)
            series.setLabelsFormat("@value")
        except Exception:
            pass

        self.chart.legend().setVisible(True)
        self.chart.legend().setAlignment(Qt.AlignBottom)

    def recalc(self):
        if len(self.lines) == 0:
            self.reset_views()
            return
        try:
            tech, eng, exp_lines, meta = self.calc()
            self.alert.hide()

            self.card_tech.set_value(str(tech.headcount), f"{tech.total_onsite_days} total days")
            self.card_eng.set_value(str(eng.headcount), f"{eng.total_onsite_days} total days")
            self.card_window.set_value(f"{meta['max_onsite']} days", f"install window {meta['window']} days")
            self.card_total.set_value(money(meta["grand_total"]), "labor + expenses")

            self.update_workload_chart(tech, eng)
            self.lbl_total_val.setText(money(meta["grand_total"]))

            rows = meta["machine_rows"]
            self.tbl_breakdown.setRowCount(len(rows))
            for r_i, r in enumerate(rows):
                # Training display rules:
                # - If training is not applicable for this model, hide all training UI/labels.
                # - If applicable but user unchecked training, show “(training excluded)”.
                if not r.get("training_applicable", True):
                    tech_disp = str(r["tech_total"])
                else:
                    if r.get("training_required", True):
                        tech_disp = f"{r['tech_total']} (incl. {r['training_days']} Train)" if r.get("training_days", 0) > 0 else str(r["tech_total"])
                    else:
                        tech_disp = f"{r['tech_total']} (training excluded)"

                if r["eng_total"] == 0:
                    eng_disp = "—"
                elif not r.get("training_applicable", True):
                    eng_disp = str(r["eng_total"])
                else:
                    eng_tp = r.get("eng_training_potential", 0)
                    eng_td = r.get("eng_training_days", 0)
                    if r.get("training_required", True):
                        eng_disp = f"{r['eng_total']} (incl. {eng_td} Train)" if (eng_tp > 0 and eng_td > 0) else str(r["eng_total"])
                    else:
                        eng_disp = f"{r['eng_total']} (training excluded)" if eng_tp > 0 else str(r["eng_total"])

                vals = [r["model"], str(r["qty"]), tech_disp, eng_disp,
                        "—" if r["tech_headcount"] == 0 else str(r["tech_headcount"]),
                        "—" if r["eng_headcount"] == 0 else str(r["eng_headcount"])]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c in [1, 4, 5]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 2 and r["training_required"]:
                        it.setForeground(Qt.darkYellow)
                    self.tbl_breakdown.setItem(r_i, c, it)

            assigns: List[Assignment] = meta["assignments"]
            self.tbl_assign.setRowCount(len(assigns))
            for i, a in enumerate(assigns):
                vals = [a.model, a.role, str(a.person_num), str(a.onsite_days), money(a.cost)]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c in [2, 3]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 4:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_assign.setItem(i, c, it)

            self.tbl_labor.setRowCount(3)
            labor_rows = [
                ("Technician", money(tech.day_rate) + "/day", str(tech.total_onsite_days), str(tech.headcount), money(tech.labor_cost)),
                ("Engineer", money(eng.day_rate) + "/day", str(eng.total_onsite_days), str(eng.headcount), money(eng.labor_cost)),
            ]
            for r_i, row in enumerate(labor_rows):
                for c, v in enumerate(row):
                    it = QTableWidgetItem(v)
                    if c in [2, 3]:
                        it.setTextAlignment(Qt.AlignCenter)
                    if c == 4:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_labor.setItem(r_i, c, it)


            # Subtotal row
            labor_subtotal = tech.labor_cost + eng.labor_cost
            self.tbl_labor.setItem(2, 0, QTableWidgetItem("Subtotal"))
            self.tbl_labor.setItem(2, 1, QTableWidgetItem(""))
            self.tbl_labor.setItem(2, 2, QTableWidgetItem(""))
            self.tbl_labor.setItem(2, 3, QTableWidgetItem(""))
            it = QTableWidgetItem(money(labor_subtotal))
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_labor.setItem(2, 4, it)

            self.lbl_exp_hdr.setText(
                f"Expenses are calculated using person-days, including {TRAVEL_DAYS_PER_PERSON} travel days per person."
            )
            self.tbl_exp.setRowCount(len(exp_lines) + 1)
            for i, l in enumerate(exp_lines):
                vals = [l.description, l.details, money(l.extended)]
                for c, v in enumerate(vals):
                    it = QTableWidgetItem(v)
                    if c == 2:
                        it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    self.tbl_exp.setItem(i, c, it)

            sub_row = len(exp_lines)
            self.tbl_exp.setItem(sub_row, 0, QTableWidgetItem("Expenses Subtotal"))
            self.tbl_exp.setItem(sub_row, 1, QTableWidgetItem("—"))
            it = QTableWidgetItem(money(meta["exp_total"]))
            it.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.tbl_exp.setItem(sub_row, 2, it)

            self.btn_print.setEnabled(True)

        except Exception as e:
            self.reset_views()
            self.alert.setText(str(e))
            self.alert.show()

    def build_quote_html(self, tech: RoleTotals, eng: RoleTotals, exp_lines: List[ExpenseLine], meta: TDict[str, object]) -> str:
        from datetime import date, timedelta
        today = date.today()
        validity = today + timedelta(days=30)
        date_str = f"{today:%B} {today.day}, {today:%Y}"
        valid_str = f"{validity:%B} {validity.day}, {validity:%Y}"

        logo_html = ""
        if LOGO_PATH.exists():
            try:
                b = LOGO_PATH.read_bytes()
                b64 = base64.b64encode(b).decode("ascii")
                logo_html = f'<img src="data:image/png;base64,{b64}" height="36" style="height:36px;" />'
            except Exception:
                logo_html = ""

        mr = []
        for r in meta["machine_rows"]:
            if not r.get("training_applicable", True):
                tech_disp = str(r["tech_total"])
            else:
                if r.get("training_required", True):
                    tech_disp = f"{r['tech_total']} (incl. {r['training_days']} Train)" if r.get("training_days", 0) > 0 else str(r["tech_total"])
                else:
                    tech_disp = f"{r['tech_total']} (training excluded)"

            if r["eng_total"] == 0:
                eng_disp = "—"
            elif not r.get("training_applicable", True):
                eng_disp = str(r["eng_total"])
            else:
                eng_tp = r.get("eng_training_potential", 0)
                eng_td = r.get("eng_training_days", 0)
                if r.get("training_required", True):
                    eng_disp = f"{r['eng_total']} (incl. {eng_td} Train)" if (eng_tp > 0 and eng_td > 0) else str(r["eng_total"])
                else:
                    eng_disp = f"{r['eng_total']} (training excluded)" if eng_tp > 0 else str(r["eng_total"])

            mr.append(f"""<tr>
                <td>{r['model']}</td>
                <td style="text-align:center;">{r['qty']}</td>
                <td>{tech_disp}</td>
                <td style="text-align:center;">{eng_disp}</td>
                <td style="text-align:center;">{r['tech_headcount'] if r['tech_headcount'] else "—"}</td>
                <td style="text-align:center;">{r['eng_headcount'] if r['eng_headcount'] else "—"}</td>
            </tr>""")

        exp_rows = []
        for l in exp_lines:
            exp_rows.append(f"""<tr>
                <td>{l.description}</td>
                <td>{l.details}</td>
                <td style="text-align:right;">{money(l.extended)}</td>
            </tr>""")

        labor_sub = tech.labor_cost + eng.labor_cost

        req_html = ""
        if self.data.requirements:
            li = "".join([f"<li>{x}</li>" for x in self.data.requirements])
            req_html = f"<h3>Requirements & Assumptions</h3><ul>{li}</ul>"

        html = f"""<html><head><meta charset="utf-8" />
        <style>
            body {{ font-family: Arial, Helvetica, sans-serif; font-size: 10pt; color: #0F172A; }}
            .topbar {{ display:flex; align-items:flex-start; justify-content:space-between; border-bottom: 3px solid #F05A28; padding-bottom: 10px; margin-bottom: 14px; }}
            .logo {{ text-align:right; }}
            .title {{ font-size: 18pt; font-weight: 800; color: #4c4b4c; margin: 0; }}
            .subtitle {{ margin: 4px 0 0 0; color: #6D6E71; }}
            .grid {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            .grid th {{ background: #343551; color: white; text-align: left; padding: 8px; border-bottom: 1px solid #E2E8F0; }}
            .grid td {{ padding: 8px; border-bottom: 1px solid #E2E8F0; }}
            .box {{ border: 1px solid #E6E8EB; border-radius: 10px; padding: 10px; background: rgba(103,144,160,0.18); }}
            .two {{ display: table; width: 100%; }}
            .two > div {{ display: table-cell; width: 50%; vertical-align: top; padding-right: 10px; }}
            h3 {{ color: #4c4b4c; margin: 18px 0 8px 0; }}
            .right {{ text-align: right; }}
            .muted {{ color: #6D6E71; }}
            .total {{ font-size: 16pt; font-weight: 900; color: #4c4b4c; }}
        </style></head><body>
            <div class="topbar">
                <div>
                    <p class="title">Commissioning Budget Quote</p>
                    <p class="subtitle muted">Service Estimate</p>
                </div>
                <div class="logo">{logo_html}</div>
            </div>

            <div class="two">
                <div class="box">
                    <b>DATE</b><br/>{date_str}<br/><br/>
                    <b>TOTAL PERSONNEL</b><br/>{tech.headcount + eng.headcount} ({tech.headcount} Tech, {eng.headcount} Eng)
                </div>
                <div class="box">
                    <b>QUOTE VALIDITY</b><br/>{valid_str}<br/><br/>
                    <b>ESTIMATED DURATION</b><br/>{meta["max_onsite"]} days onsite + {TRAVEL_DAYS_PER_PERSON} travel days
                </div>
            </div>
            <div class="section-spacer"></div>

            <h3>Machine Breakdown</h3>
            <table class="grid">
                <tr><th>Model</th><th style="text-align:center;">Qty</th><th>Tech Days</th><th style="text-align:center;">Eng Days</th>
                    <th style="text-align:center;">Technicians</th><th style="text-align:center;">Engineers</th></tr>
                {''.join(mr)}
            </table>

            <h3>Labor Costs</h3>
            <table class="grid">
                <tr><th>Item</th><th class="right">Extended</th></tr>
                <tr><td>Tech. Regular Time ({tech.total_onsite_days} days × {money(tech.day_rate)}/day)</td><td class="right">{money(tech.labor_cost)}</td></tr>
                <tr><td>Eng. Regular Time ({eng.total_onsite_days} days × {money(eng.day_rate)}/day)</td><td class="right">{money(eng.labor_cost)}</td></tr>
                <tr><td><b>Labor Subtotal</b></td><td class="right"><b>{money(labor_sub)}</b></td></tr>
            </table>

            <h3>Estimated Expenses</h3>
            <div class="muted">Includes {int(meta["total_trip_days"])} total trip day(s) across personnel (onsite + travel days).</div>
            <table class="grid">
                <tr><th>Expense</th><th>Details</th><th class="right">Amount</th></tr>
                {''.join(exp_rows)}
                <tr><td><b>Expenses Subtotal</b></td><td>—</td><td class="right"><b>{money(meta["exp_total"])}</b></td></tr>
            </table>

            <h3>Estimated Total</h3>
            <div class="box">
                <span class="total">{money(meta["grand_total"])}</span><br/>
                <span class="muted">Labor ({money(labor_sub)}) + Expenses ({money(meta["exp_total"])})</span>
            </div>

            <h3>Terms & Conditions</h3>
            <ul>
                <li><b>Pricing & Quote Expiration:</b> Prices shown reflect an estimate of days and expenses. Any additional time will be billed at the rates shown. Quote valid for 30 days.</li>
                <li><b>Customer Install Window:</b> No individual technician or engineer is assigned more than {meta["window"]} onsite days per trip.</li>
                <li><b>Training:</b> Training days are calculated at 1 day per {TRAINING_MACHINES_PER_DAY} machines of the same model type. Training can be excluded per machine if not required (customer request only).</li>
                <li><b>Machine-Specific Skills:</b> Each machine type requires technicians with specialized skills. Personnel are not shared across different machine types.</li>
                <li><b>Travel Days:</b> Expenses include {TRAVEL_DAYS_PER_PERSON} travel days (1 day travel-in + 1 day travel-out) in addition to onsite work days.</li>
            </ul>
            {req_html}
        </body></html>"""
        return html

    def print_quote_preview(self):
        try:
            tech, eng, exp_lines, meta = self.calc()
        except Exception as e:
            QMessageBox.critical(self, "Cannot print", str(e))
            return

        try:
            html = self.build_quote_html(tech, eng, exp_lines, meta)
            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter(QPrinter.HighResolution)
            # Letter page (8.5x11)
            try:
                printer.setPageSize(QPageSize(QPageSize.Letter))
            except Exception:
                pass

            preview = QPrintPreviewDialog(printer, self)
            preview.setWindowTitle("Print Preview - Commissioning Budget Quote")
            preview.setWindowModality(Qt.ApplicationModal)
            preview.resize(1100, 800)
            preview.paintRequested.connect(lambda p: doc.print_(p))

            # Show the dialog reliably
            preview.exec()
        except Exception as e:
            QMessageBox.critical(self, "Print error", str(e))
            return

    
    def _update_right_scroll_height_if_stacked(self):
        """When stacked, expand the right scroll area to its content so the OUTER scroll handles scrolling."""
        if not getattr(self, "_is_stacked", False):
            return
        try:
            if hasattr(self, "right_scroll") and hasattr(self, "right_content"):
                h = int(self.right_content.sizeHint().height()) + 80
                self.right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
                self.right_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                self.right_scroll.setMinimumHeight(h)
                self.right_scroll.setMaximumHeight(h)
        except Exception:
            pass

    def _apply_responsive_layout(self):
        """Large screens: 2 columns w/ right-side scroll. Small screens: single stacked w/ full-window scroll."""
        if not hasattr(self, "splitter") or not hasattr(self, "outer_scroll") or not hasattr(self, "right_scroll"):
            return

        w = int(self.width())
        stacked = w < getattr(self, "_stack_threshold", 1280)

        if stacked and not getattr(self, "_is_stacked", False):
            self._is_stacked = True
            self.splitter.setOrientation(Qt.Vertical)
            # Enable whole-window scrolling; disable inner right scrolling
            self.outer_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.outer_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self._update_right_scroll_height_if_stacked()
            # Give left pane enough room; right pane will follow under it
            try:
                self.splitter.setSizes([650, 1000])
            except Exception:
                pass

            # In stacked (single-column) mode, make the machine configuration area taller so
            # multiple machine lines are visible without feeling cramped.
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(320)
                    self.scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            except Exception:
                pass

        elif (not stacked) and getattr(self, "_is_stacked", False):
            self._is_stacked = False
            self.splitter.setOrientation(Qt.Horizontal)
            # Disable whole-window scrolling; allow right column to scroll
            self.outer_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.outer_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.right_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.right_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            self.right_scroll.setMinimumHeight(0)
            self.right_scroll.setMaximumHeight(16777215)
            self.right_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            try:
                self.splitter.setSizes([520, 1040])
            except Exception:
                pass

            # Restore default sizing for wide (two-column) mode.
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(0)
            except Exception:
                pass

        elif stacked:
            # Still stacked; keep heights updated as content changes
            self._update_right_scroll_height_if_stacked()
            try:
                if hasattr(self, "scroll"):
                    self.scroll.setMinimumHeight(320)
            except Exception:
                pass

    def _apply_scale(self):
        # Scale UI typography modestly with window size; keep within sensible bounds.
        w = max(self.width(), 1)
        h = max(self.height(), 1)
        # Use width as primary driver; clamp to avoid extremes.
        scale = w / 1920.0
        scale = 0.85 if scale < 0.85 else (1.25 if scale > 1.25 else scale)
        pt = self._base_font_pt * scale
        f = QFont(self.font())
        f.setPointSizeF(pt)
        self.setFont(f)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._apply_responsive_layout()
        self._apply_scale()
    def closeEvent(self, event):

        event.accept()


def main():
    app = QApplication(sys.argv)
    w = MainWindow()
    w.showMaximized()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
